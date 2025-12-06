#!/usr/bin/env python3
"""
Create Excel File with Dynamic Formulas

This script converts the properties CSV to an Excel file with dynamic formulas
that automatically calculate scores based on price, distance to Chinatown, and size.
The formulas update automatically when you edit values in Excel.
"""

import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_column_letter(col_num):
    """Convert column number to Excel letter (1=A, 2=B, etc.)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + 65) + result
        col_num //= 26
    return result


def add_formulas_to_excel(excel_file, price_col, distance_col, size_col, data_start_row, data_end_row):
    """
    Add dynamic formulas to the Excel file for calculating scores.

    Args:
        excel_file: Path to Excel file
        price_col: Column letter for price (e.g., 'B')
        distance_col: Column letter for distance (e.g., 'X')
        size_col: Column letter for size (e.g., 'Y')
        data_start_row: First row with data (usually 2, after header)
        data_end_row: Last row with data
    """
    logger.info("Adding formulas to Excel file...")

    # Load the workbook
    wb = load_workbook(excel_file)
    ws = wb.active

    # Find where to add the new columns (after existing columns)
    max_col = ws.max_column

    # Add headers for score columns
    score_start_col = max_col + 1
    headers = {
        score_start_col: 'price_score',
        score_start_col + 1: 'distance_score',
        score_start_col + 2: 'size_score',
        score_start_col + 3: 'combined_score'
    }

    for col_num, header in headers.items():
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Get column letters for score columns
    price_score_col = get_column_letter(score_start_col)
    distance_score_col = get_column_letter(score_start_col + 1)
    size_score_col = get_column_letter(score_start_col + 2)
    combined_score_col = get_column_letter(score_start_col + 3)

    logger.info(f"Adding formulas from row {data_start_row} to {data_end_row}...")

    # Add formulas for each data row
    for row in range(data_start_row, data_end_row + 1):
        price_cell = f"{price_col}{row}"
        distance_cell = f"{distance_col}{row}"
        size_cell = f"{size_col}{row}"

        # Price Score Formula: Normalize price (lower is better)
        # Formula: IF(AND(price, min, max exist), (MAX-price)/(MAX-MIN)*100, "")
        price_score_formula = (
            f'=IF(AND(ISNUMBER({price_cell}), '
            f'ISNUMBER(MIN({price_col}${data_start_row}:{price_col}${data_end_row})), '
            f'ISNUMBER(MAX({price_col}${data_start_row}:{price_col}${data_end_row}))), '
            f'IF(MAX({price_col}${data_start_row}:{price_col}${data_end_row})=MIN({price_col}${data_start_row}:{price_col}${data_end_row}), 50, '
            f'ROUND((MAX({price_col}${data_start_row}:{price_col}${data_end_row})-{price_cell})/'
            f'(MAX({price_col}${data_start_row}:{price_col}${data_end_row})-MIN({price_col}${data_start_row}:{price_col}${data_end_row}))*100, 2)), "")'
        )

        # Distance Score Formula: Normalize distance (closer is better)
        distance_score_formula = (
            f'=IF(AND(ISNUMBER({distance_cell}), '
            f'ISNUMBER(MIN({distance_col}${data_start_row}:{distance_col}${data_end_row})), '
            f'ISNUMBER(MAX({distance_col}${data_start_row}:{distance_col}${data_end_row}))), '
            f'IF(MAX({distance_col}${data_start_row}:{distance_col}${data_end_row})=MIN({distance_col}${data_start_row}:{distance_col}${data_end_row}), 50, '
            f'ROUND((MAX({distance_col}${data_start_row}:{distance_col}${data_end_row})-{distance_cell})/'
            f'(MAX({distance_col}${data_start_row}:{distance_col}${data_end_row})-MIN({distance_col}${data_start_row}:{distance_col}${data_end_row}))*100, 2)), "")'
        )

        # Size Score Formula: Normalize size (larger is better)
        size_score_formula = (
            f'=IF(AND(ISNUMBER({size_cell}), '
            f'ISNUMBER(MIN({size_col}${data_start_row}:{size_col}${data_end_row})), '
            f'ISNUMBER(MAX({size_col}${data_start_row}:{size_col}${data_end_row}))), '
            f'IF(MAX({size_col}${data_start_row}:{size_col}${data_end_row})=MIN({size_col}${data_start_row}:{size_col}${data_end_row}), 50, '
            f'ROUND(({size_cell}-MIN({size_col}${data_start_row}:{size_col}${data_end_row}))/'
            f'(MAX({size_col}${data_start_row}:{size_col}${data_end_row})-MIN({size_col}${data_start_row}:{size_col}${data_end_row}))*100, 2)), "")'
        )

        # Combined Score Formula: Average of three scores (equal weighting)
        combined_score_formula = (
            f'=IF(AND(ISNUMBER({price_score_col}{row}), ISNUMBER({distance_score_col}{row}), ISNUMBER({size_score_col}{row})), '
            f'ROUND(({price_score_col}{row}+{distance_score_col}{row}+{size_score_col}{row})/3, 2), "")'
        )

        # Add formulas to cells
        ws[f"{price_score_col}{row}"] = price_score_formula
        ws[f"{distance_score_col}{row}"] = distance_score_formula
        ws[f"{size_score_col}{row}"] = size_score_formula
        ws[f"{combined_score_col}{row}"] = combined_score_formula

    logger.info("✓ Formulas added successfully")

    # Add conditional formatting (color scale) to combined_score column
    logger.info("Adding conditional formatting...")

    # Color scale: Red (low) -> Yellow (mid) -> Green (high)
    combined_score_range = f"{combined_score_col}{data_start_row}:{combined_score_col}{data_end_row}"
    ws.conditional_formatting.add(
        combined_score_range,
        ColorScaleRule(
            start_type='min', start_color='F8696B',  # Red
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
            end_type='max', end_color='63BE7B'  # Green
        )
    )

    logger.info("✓ Conditional formatting added")

    # Adjust column widths
    logger.info("Adjusting column widths...")
    ws.column_dimensions[price_score_col].width = 12
    ws.column_dimensions[distance_score_col].width = 15
    ws.column_dimensions[size_score_col].width = 12
    ws.column_dimensions[combined_score_col].width = 15

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save the workbook
    logger.info("Saving Excel file...")
    wb.save(excel_file)
    logger.info("✓ Excel file saved with formulas")


def main():
    """Main function to create Excel file with formulas."""

    # File paths
    input_file = "/Users/simeontu/Documents/vscode-projects/rightmove_webscraper.py/results/enhanced_scrape_2025-Oct-16_at_17h42m/enhanced_properties_with_chinatown_distance.csv"
    output_file = "/Users/simeontu/Documents/vscode-projects/rightmove_webscraper.py/results/enhanced_scrape_2025-Oct-16_at_17h42m/enhanced_properties_dynamic.xlsx"

    logger.info("=" * 80)
    logger.info("CREATE EXCEL FILE WITH DYNAMIC FORMULAS")
    logger.info("=" * 80)
    logger.info(f"Input file: {input_file}")
    logger.info(f"Output file: {output_file}")
    logger.info("=" * 80)

    # Load the CSV
    try:
        df = pd.read_csv(input_file)
        logger.info(f"✓ Loaded {len(df)} properties")
    except Exception as e:
        logger.error(f"✗ Error loading CSV: {e}")
        return 1

    # Check required columns exist
    required_columns = ['price', 'distance_to_chinatown_km', 'property_size_sqm']
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        logger.error(f"✗ Missing required columns: {missing_columns}")
        return 1

    # Get column positions (Excel is 1-indexed)
    price_col_idx = df.columns.get_loc('price') + 1
    distance_col_idx = df.columns.get_loc('distance_to_chinatown_km') + 1
    size_col_idx = df.columns.get_loc('property_size_sqm') + 1

    price_col = get_column_letter(price_col_idx)
    distance_col = get_column_letter(distance_col_idx)
    size_col = get_column_letter(size_col_idx)

    logger.info(f"✓ Column mappings:")
    logger.info(f"  Price: Column {price_col}")
    logger.info(f"  Distance: Column {distance_col}")
    logger.info(f"  Size: Column {size_col}")

    # Export to Excel first (without formulas)
    logger.info(f"\nExporting to Excel...")
    try:
        df.to_excel(output_file, index=False, engine='openpyxl')
        logger.info("✓ Excel file created")
    except Exception as e:
        logger.error(f"✗ Error creating Excel file: {e}")
        return 1

    # Add formulas to the Excel file
    data_start_row = 2  # Row 1 is header, data starts at row 2
    data_end_row = len(df) + 1  # Last row with data

    try:
        add_formulas_to_excel(
            output_file,
            price_col,
            distance_col,
            size_col,
            data_start_row,
            data_end_row
        )
    except Exception as e:
        logger.error(f"✗ Error adding formulas: {e}")
        return 1

    # Display summary
    properties_with_complete_data = df[
        df['price'].notna() &
        df['distance_to_chinatown_km'].notna() &
        df['property_size_sqm'].notna()
    ]

    logger.info("\n" + "=" * 80)
    logger.info("SUMMARY")
    logger.info("=" * 80)
    logger.info(f"Total properties: {len(df)}")
    logger.info(f"Properties with complete data: {len(properties_with_complete_data)} ({len(properties_with_complete_data)/len(df)*100:.1f}%)")
    logger.info(f"\nThe Excel file contains dynamic formulas that will:")
    logger.info(f"  • Automatically calculate scores when you edit values")
    logger.info(f"  • Update combined_score based on price, distance, and size")
    logger.info(f"  • Use equal weighting (33.33% each)")
    logger.info(f"  • Show color gradient (red=low, yellow=mid, green=high)")
    logger.info("\n" + "=" * 80)
    logger.info("COMPLETE!")
    logger.info("=" * 80)
    logger.info(f"Output file: {output_file}")
    logger.info("=" * 80)

    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
